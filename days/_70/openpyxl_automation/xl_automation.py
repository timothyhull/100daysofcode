#!/usr/bin/env python3

""" openpyxl automation script for day 70 of 100DaysOfCode. """

# Imports - Python Standard Library
from os import path
from pathlib import Path

# Imports - Third-Party
from openpyxl import load_workbook, Workbook

# Imports - Local

# Constants
FILE_NAME = path.abspath(__file__)
LOCAL_DIR = Path(FILE_NAME).parent
DATA_DIR = 'data'
DATA_FILE = 'Financial_Sample.xlsx'
DATA_FILE_PATH = path.join(LOCAL_DIR, DATA_DIR, DATA_FILE)


# Read the spreadsheet data
def import_workbook() -> None:
    """ Import Excel workbook.

        Args:
            None.

        Returns:
            None.
    """

    # Load the workbook file
    print(f'\nLoading workbook "{DATA_FILE}"...', end='')
    workbook = load_workbook(
        filename=DATA_FILE_PATH
    )
    print('done.\n')

    # Display worksheet names
    print('Worksheet Names:')
    for index, worksheet in enumerate(workbook.sheetnames, 1):
        print(f'{index}. {worksheet}')

    return workbook


def get_profit_total(
    workbook: Workbook
) -> float:
    """ Get a total of all profits from a workbook.

        Args:
            workbook (openpyxl.Workbook):
                openpyxl Workbook object, created using the
                load_workbook method of openpyxl.

        Returns:
            profit_total (float):
                Total of all cells in the Profits Column.
    """

    # Set profit_total to 0
    profit_total = 0

    # Get the first worksheet from the spreadsheet
    worksheet_1 = workbook[workbook.sheetnames[0]]

    # Set the column of values
    column = 'L'

    # Create a list of all column cells
    # Loop over all columns in the worksheet
    for col in worksheet_1:
        # Loop over a range of row numbers
        # for row in range(2, 102):
        # Replace the specific range end number with the max_row value
        for row in range(2, worksheet_1.max_row):
            """ Set the current cell to a string value of the column name (L)
            plus the current row number. """
            current_cell = f'{column}{row}'

            # Conditionally add the value of the current cell to profit_total
            if worksheet_1[current_cell].value is not None:
                profit_total += float(f'{worksheet_1[current_cell].value}')

    return round(profit_total, 2)


def main() -> None:
    """ Main program. """

    # Import the workbook file
    workbook = import_workbook()

    # Get a total of all profits from the first worksheet in the workbook
    profit_total = get_profit_total(workbook)
    print(f'Total profits: ${profit_total}')

    return None


if __name__ == '__main__':
    main()
